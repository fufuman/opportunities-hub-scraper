import argparse
import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

UNIFIED_COLUMNS = [
    "school_name",
    "faculty_name",
    "title",
    "department",
    "medium",
    "email",
    "email_type",
    "source_url",
    "date_extracted",
]

# One citation per school: where its faculty data was publicly listed, and any
# notable caveats about how medium/department was determined. Update this
# whenever a school's scraper is added or its source pages change.
SOURCE_CITATIONS = {
    "Temple/Tyler": {
        "name": "Temple University Tyler School of Art and Architecture - full "
                "directory (Staff/Faculty/Adjunct Faculty filter), 10 pages "
                "(source brief overclaimed 16 pages; page controls confirm 10). "
                "JS-rendered - required crawl4ai, not a plain fetch.",
        "url": "https://tyler.temple.edu/directory?profile-type%5Bstaff%5D=staff&"
               "profile-type%5Bfaculty%5D=faculty&profile-type%5Badjunct_faculty%5D="
               "adjunct_faculty&page=0 (through page=9)",
        "accessed": "2026-08-08",
        "country": "United States",
        "city": "Philadelphia, PA",
    },
    "Ohio State": {
        "name": "Ohio State University Department of Art - /people directory "
                "(Chair/Faculty/Associated Faculty/Emeritus Faculty categories only, "
                "Staff and Graduate Students excluded) cross-referenced with each "
                "person's individual profile page bio text for medium/discipline, "
                "since the directory listing itself has no per-person area label "
                "(contrary to the research brief) and the /areas/<medium> pages have "
                "no faculty roster at all (descriptive copy only).",
        "url": "https://art.osu.edu/people (plus 61 individual /people/<id> profile pages)",
        "accessed": "2026-08-08",
        "country": "United States",
        "city": "Columbus, OH",
    },
    "UGA": {
        "name": "University of Georgia, Lamar Dodd School of Art - full directory, "
                "10 pages, medium classified from the 'Academic Area' field shown "
                "directly on each person's card.",
        "url": "https://art.uga.edu/directory/ (page/2/ through page/10/)",
        "accessed": "2026-08-09",
        "country": "United States",
        "city": "Athens, GA",
    },
    "University of Iowa": {
        "name": "University of Iowa, School of Art, Art History and Design - "
                "/people/faculty directory, 2 pages, medium classified from the "
                "Title/Position field(s) shown directly on each person's card "
                "(e.g. 'Professor of Sculpture & Intermedia').",
        "url": "https://art.uiowa.edu/people/faculty (and ?page=1)",
        "accessed": "2026-08-09",
        "country": "United States",
        "city": "Iowa City, IA",
    },
    "UT Austin": {
        "name": "University of Texas at Austin, Department of Art and Art History - "
                "/about/who-we-are/people directory (Faculty, Staff & Students "
                "combined listing), 5 pages, medium classified from the "
                "'Designation' field (e.g. 'Studio Art (Painting & Drawing)'). "
                "Emails were NOT obfuscated with spaces as the research brief "
                "claimed - they render plainly, only wrapped with <wbr> tags for "
                "line-breaking, which were stripped.",
        "url": "https://art.utexas.edu/about/who-we-are/people (and ?page=1 through ?page=4)",
        "accessed": "2026-08-09",
        "country": "United States",
        "city": "Austin, TX",
    },
    "RIT": {
        "name": "Rochester Institute of Technology, College of Art and Design - "
                "20 per-program faculty directory pages (program-directory/<id>), "
                "IDs found via each program's official landing page rather than "
                "guessed. Covers all in-scope BFA/MFA programs (Art Education MST "
                "excluded); Furniture Design AOS, Photographic Arts Exploration, "
                "and Studio Arts Exploration have no faculty-directory link and "
                "were skipped. Medium assigned per-program (e.g. Ceramics MFA -> "
                "Sculpture), not from a per-person field. Faculty cross-listed on "
                "multiple program pages (e.g. BFA + MFA in the same area) "
                "deduped, keeping the first page's medium assignment.",
        "url": "https://www.rit.edu/artdesign/program-directory/411480 (and 19 more IDs)",
        "accessed": "2026-08-09",
        "country": "United States",
        "city": "Rochester, NY",
    },
    "Edinburgh": {
        "name": "Edinburgh College of Art, University of Edinburgh - /people "
                "directory filtered server-side to Academic Staff + Key Academic "
                "Office Holders only (excludes Honorary/Emeritus, Postgraduate "
                "Research Students, Professional Services, Student "
                "Representation), 13 pages. Medium classified from the role/title "
                "text shown directly on each card.",
        "url": "https://www.eca.ed.ac.uk/people?field_people_type_target_id%5B17%5D="
               "17&field_people_type_target_id%5B16%5D=16&page=0 (through page=12)",
        "accessed": "2026-08-09",
        "country": "United Kingdom",
        "city": "Edinburgh, Scotland",
    },
    "SNU": {
        "name": "Seoul National University, College of Fine Arts - 5 per-department "
                "faculty pages (Design, Painting, Sculpture, Craft, Oriental "
                "Painting). JS-rendered - required crawl4ai. Medium classified "
                "from each person's research-area text, falling back to the "
                "department itself when the area text has no medium keyword "
                "(e.g. 'Theory of Art').",
        "url": "https://art.snu.ac.kr/en/category/design-en/?catemenu=Faculty&type=major "
               "(and 4 more department category pages)",
        "accessed": "2026-08-09",
        "country": "South Korea",
        "city": "Seoul",
    },
    "VCU": {
        "name": "Virginia Commonwealth University, School of the Arts (VCUarts) - "
                "shared /directory/ filtered per-department via query param, 8 "
                "in-scope departments (Painting + Printmaking, Kinetic Imaging, "
                "Sculpture + Extended Media, Photography + Film, Communication "
                "Arts, Graphic Design, Craft and Material Studies, Fashion Design "
                "and Merchandising). Department pages mix teaching faculty with "
                "administrative/support staff under the same listing; rows with "
                "a clearly non-teaching title (Coordinator, Administrative, "
                "Manager, Technician, Advisor) were excluded.",
        "url": "https://arts.vcu.edu/directory/?department%5B%5D=painting-printmaking "
               "(and 7 more department query values)",
        "accessed": "2026-08-09",
        "country": "United States",
        "city": "Richmond, VA",
    },
    "NID Ahmedabad": {
        "name": "National Institute of Design, Ahmedabad - /people/faculty page. "
                "NOTE: this page has no pagination and lists only 9 faculty total "
                "(Industrial Design, Textile & Apparel Design, Communication "
                "Design) - Ceramic/Glass and Film/Animation faculty mentioned in "
                "the research brief were NOT found; confirmed via the site's own "
                "academic program pages that Ceramic and Glass Design and "
                "Animation Film Design are not current B.Des specializations at "
                "this campus (may exist at a different NID campus, or may have "
                "been discontinued/restructured). This may be the complete "
                "current roster shown on the public site, not a scraper gap.",
        "url": "https://www.nid.ac.in/people/faculty",
        "accessed": "2026-08-09",
        "country": "India",
        "city": "Ahmedabad",
    },
    "Jamia Millia Islamia": {
        "name": "Jamia Millia Islamia, Faculty of Fine Arts - per-department "
                "'Faculty Members' pages (Painting, Sculpture, Applied Art). "
                "These pages are JS-driven search widgets that return no static "
                "data on a plain fetch, but auto-load that department's own "
                "roster by default when rendered - required crawl4ai. The "
                "university-wide faculty search and the FFA-level staff search "
                "page do NOT auto-load results and were not usable. Art "
                "Education and Art History departments excluded (out of scope); "
                "Graphic Art department not attempted this pass.",
        "url": "https://jmi.ac.in/ACADEMICS/Departments/Department-Of-Painting/Faculty-Members "
               "(and Department-Of-Sculpture, Department-Of-Applied-Art)",
        "accessed": "2026-08-09",
        "country": "India",
        "city": "New Delhi",
    },
    "UIC": {
        "name": "University of Illinois Chicago, School of Art and Art History - "
                "/content/art-faculty studio listing page (Art History faculty "
                "excluded per user's medium scope), split into sections (Studio "
                "Arts, Photography, New Media Arts, Moving Image, Interdisciplinary "
                "Degree in the Arts; Art Education section excluded). Names link "
                "to individual profile pages where email is the only place it's "
                "published - required a profile click-through pass, same pattern "
                "as Ohio State. Emeriti section has no profile links at all (plain "
                "text names only) and could not be scraped this way - skipped.",
        "url": "https://artandarthistory.uic.edu/content/art-faculty "
               "(plus 13 individual /profile/<slug> pages)",
        "accessed": "2026-08-09",
        "country": "United States",
        "city": "Chicago, IL",
    },
    "Rutgers": {
        "name": "Rutgers University, Mason Gross School of the Arts - Art & Design "
                "department faculty/staff page. Emails ARE visible directly on the "
                "listing (better than the research brief's Tier 2 classification "
                "implied) - no profile click-through needed. Only faculty whose "
                "title text names a specific medium were kept (e.g. 'Associate "
                "Professor in Photography'); generic titles ('Professor') with no "
                "medium signal were left out rather than guessed, since this page "
                "has no separate discipline field. One email format quirk: a few "
                "hrefs are 'mailto:Name <email>' URL-encoded rather than a bare "
                "address - the actual address was extracted from either form.",
        "url": "https://www.masongross.rutgers.edu/degrees-programs/art-design/faculty-staff/",
        "accessed": "2026-08-09",
        "country": "United States",
        "city": "New Brunswick, NJ",
    },
    "Cornell": {
        "name": "Cornell University, Department of Art (AAP) - /art/art-people "
                "listing (41 entries, 5 duplicates from dual program listings, "
                "deduped to 36 unique profiles) with emails only on individual "
                "profile pages - required a click-through pass like Ohio State. "
                "Medium classified from title/role text first, falling back to "
                "each profile's bio paragraph when the role alone was generic "
                "(e.g. 'Professor'). 5 rows have email_type=department_general: "
                "several visiting/affiliated faculty list the shared "
                "imagetext@cornell.edu program mailbox as their own contact on "
                "their profile page, not a personal address - flagged rather than "
                "treated as equivalent to a direct personal email.",
        "url": "https://aap.cornell.edu/art/art-people (plus 36 individual /people/<slug>/ pages)",
        "accessed": "2026-08-09",
        "country": "United States",
        "city": "Ithaca, NY",
    },
    "Syracuse": {
        "name": "Syracuse University, College of Visual and Performing Arts, School "
                "of Art - /academics/art/contact/ page (the master faculty-staff "
                "directory is a JS search widget with no default results, "
                "confirmed unusable per the brief). This page has the School of "
                "Art Director (medium classified from bio text on his own profile "
                "page - no discipline given on the contact page for this role) "
                "plus 4 'Area Leads' with a discipline label given directly "
                "(e.g. 'Printmaking', 'Illustration B.F.A.') but email only on "
                "their individual profile page. 1 area lead (Studio Arts B.F.A./"
                "B.S., generalist label with no specific medium) excluded rather "
                "than guessed.",
        "url": "https://vpa.syracuse.edu/academics/art/contact/ (plus 4 individual "
               "/faculty-staff/<slug>/ profile pages)",
        "accessed": "2026-08-09",
        "country": "United States",
        "city": "Syracuse, NY",
    },
    "Slade": {
        "name": "Slade School of Fine Art, University College London (UCL) - "
                "/people/academic/ listing (46 people, single page, no "
                "pagination). Medium classified from title text shown directly "
                "on the listing (e.g. 'Lecturer, Painting', 'Associate Professor, "
                "Sculpture'); email only on individual profile pages - required a "
                "click-through pass for the 21 in-scope people.",
        "url": "https://www.ucl.ac.uk/slade/people/academic/ (plus 21 individual profile pages)",
        "accessed": "2026-08-09",
        "country": "United Kingdom",
        "city": "London, England",
    },
    "Yale": {
        "name": "Yale School of Art - the brief correctly found the "
                "/about/people/faculty-and-staff listing page has no emails, "
                "classifying Yale as Tier 3. However each of the 74 people "
                "listed (Academic Leadership, Graphic Design, Painting/"
                "Printmaking, Photography, Sculpture, Interdepartmental, "
                "Undergraduate, Yale Norfolk, Faculty Emeriti sections; Faculty "
                "Governing Board and Administration/Staff excluded) links to "
                "their own profile page, and SOME publish a direct email there "
                "(appears to be individual choice - several senior faculty link "
                "to a personal site/Substack instead). Only 6 of 49 in-scope "
                "profiles checked had a public email.",
        "url": "https://www.art.yale.edu/about/people/faculty-and-staff "
               "(plus 49 individual /<Name> profile pages)",
        "accessed": "2026-08-09",
        "country": "United States",
        "city": "New Haven, CT",
    },
    "CMU": {
        "name": "Carnegie Mellon University, School of Art - the brief's "
                "warning (\"directory requires Andrew login, do NOT scrape\") "
                "correctly applies to CMU's university-wide people directory, "
                "which was NOT used here. This is a different, fully public "
                "page - art.cmu.edu/people/faculty/ - the School of Art's own "
                "faculty listing, not login-walled. Individual profile pages "
                "publish a real institutional email (@andrew.cmu.edu or @cmu.edu) "
                "in a structured JSON metadata field. Medium classified from the "
                "discipline label shown directly on the listing (e.g. 'Drawing, "
                "Painting, Print, & Photo', 'Electronic & Time Based Media' - "
                "the latter mapped to Filmmaking as the closest of the 9 "
                "in-scope mediums, though it also covers interactive/digital "
                "work; an approximate call).",
        "url": "https://art.cmu.edu/people/faculty/ (plus 14 individual /people/<slug>/ pages)",
        "accessed": "2026-08-09",
        "country": "United States",
        "city": "Pittsburgh, PA",
    },
    "GSA": {
        "name": "Glasgow School of Art - the brief correctly found the /staff "
                "listing page has no emails, classifying GSA as Tier 3. However "
                "individual profile pages DO publish a real email (e.g. "
                "'Email: N.Oddy@gsa.ac.uk') even though the listing itself "
                "doesn't show it. ~197 total staff across all departments "
                "(architecture, design history, admin, etc., 2 pages); medium "
                "classified from job title text, keeping only the ~29 whose "
                "title names an in-scope medium (e.g. 'Lecturer in Painting and "
                "Printmaking', 'Lecturer Sculpture and Environmental Art').",
        "url": "https://www.gsa.ac.uk/staff (2 pages, plus ~29 individual /staff/user-<id> pages)",
        "accessed": "2026-08-09",
        "country": "United Kingdom",
        "city": "Glasgow, Scotland",
    },
    "Kunstakademie Dusseldorf": {
        "name": "Kunstakademie Dusseldorf - the brief's Tier 3 classification "
                "(no emails, use postmaster@) is correct for the vast majority: "
                "checked all 24 'Freie Kunst' (Free Art) professors' individual "
                "profile pages (JS-rendered, required crawl4ai) and only 1 "
                "(Alexandra Bircken) publishes an email, included in her "
                "application-instructions text ('send your portfolio to "
                "alexandra.bircken@kunstakademie-duesseldorf.de'). The other 23 "
                "genuinely have no email anywhere on their page, confirmed by "
                "direct inspection, not a fetch failure. Baukunst (Architecture) "
                "and Kunstbezogene Wissenschaften (art theory/sciences) sections "
                "excluded as out of scope.",
        "url": "https://kunstakademie-duesseldorf.de/studienangebot-und-bewerbung/professor-innen/ "
               "(plus 24 individual profile pages)",
        "accessed": "2026-08-09",
        "country": "Germany",
        "city": "Dusseldorf",
    },
}


def read_csv_rows(csv_path):
    with open(csv_path, "r", newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def write_sheet(wb, sheet_title, rows, school_label):
    ws = wb.create_sheet(title=sheet_title[:31])

    citation = SOURCE_CITATIONS.get(school_label)
    header_row_offset = 0
    if citation:
        ws.append([f"Country: {citation['country']}"])
        ws.append([f"City: {citation['city']}"])
        ws.append([f"Source: {citation['name']}"])
        ws.append([f"URL: {citation['url']}"])
        ws.append([f"Accessed: {citation['accessed']}"])
        for r in range(1, 6):
            ws.cell(row=r, column=1).font = Font(italic=True, size=9, color="666666")
        ws.append([])
        header_row_offset = 6

    header_row_num = header_row_offset + 1
    ws.append(UNIFIED_COLUMNS)
    for cell in ws[header_row_num]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([row.get(col, "") for col in UNIFIED_COLUMNS])

    last_row = ws.max_row
    if last_row <= header_row_num:
        last_row = header_row_num + 1
    first_data_row = header_row_num + 1

    ws.freeze_panes = f"A{first_data_row}"
    ws.auto_filter.ref = f"A{header_row_num}:{get_column_letter(len(UNIFIED_COLUMNS))}{last_row}"

    widths = {"school_name": 34, "faculty_name": 22, "title": 30, "department": 22,
              "medium": 20, "email": 30, "email_type": 14, "source_url": 40,
              "date_extracted": 16}
    for i, col in enumerate(UNIFIED_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 18)

    return ws


def main():
    parser = argparse.ArgumentParser(description="Build/update the master faculty outreach workbook")
    parser.add_argument("--out", default="master_faculty.xlsx")
    parser.add_argument("--temple-csv", default="temple_faculty.csv")
    parser.add_argument("--ohio-state-csv", default="ohio_state_faculty.csv")
    parser.add_argument("--uga-csv", default="uga_faculty.csv")
    parser.add_argument("--iowa-csv", default="iowa_faculty.csv")
    parser.add_argument("--ut-austin-csv", default="ut_austin_faculty.csv")
    parser.add_argument("--rit-csv", default="rit_faculty.csv")
    parser.add_argument("--edinburgh-csv", default="edinburgh_faculty.csv")
    parser.add_argument("--snu-csv", default="snu_faculty.csv")
    parser.add_argument("--vcu-csv", default="vcu_faculty.csv")
    parser.add_argument("--nid-csv", default="nid_faculty.csv")
    parser.add_argument("--jamia-csv", default="jamia_faculty.csv")
    parser.add_argument("--uic-csv", default="uic_faculty.csv")
    parser.add_argument("--rutgers-csv", default="rutgers_faculty.csv")
    parser.add_argument("--cornell-csv", default="cornell_faculty.csv")
    parser.add_argument("--syracuse-csv", default="syracuse_faculty.csv")
    parser.add_argument("--slade-csv", default="slade_faculty.csv")
    parser.add_argument("--yale-csv", default="yale_faculty.csv")
    parser.add_argument("--cmu-csv", default="cmu_faculty.csv")
    parser.add_argument("--gsa-csv", default="gsa_faculty.csv")
    parser.add_argument("--duesseldorf-csv", default="duesseldorf_faculty.csv")
    args = parser.parse_args()

    wb = Workbook()
    wb.remove(wb.active)

    # (sheet title, csv path, school label). Add a new tuple here for each
    # additional school scraped; the CSV just needs the unified columns.
    sources = [
        ("Temple-Tyler", args.temple_csv, "Temple/Tyler"),
        ("Ohio State", args.ohio_state_csv, "Ohio State"),
        ("UGA", args.uga_csv, "UGA"),
        ("Iowa", args.iowa_csv, "University of Iowa"),
        ("UT Austin", args.ut_austin_csv, "UT Austin"),
        ("RIT", args.rit_csv, "RIT"),
        ("Edinburgh", args.edinburgh_csv, "Edinburgh"),
        ("SNU", args.snu_csv, "SNU"),
        ("VCU", args.vcu_csv, "VCU"),
        ("NID Ahmedabad", args.nid_csv, "NID Ahmedabad"),
        ("Jamia Millia Islamia", args.jamia_csv, "Jamia Millia Islamia"),
        ("UIC", args.uic_csv, "UIC"),
        ("Rutgers", args.rutgers_csv, "Rutgers"),
        ("Cornell", args.cornell_csv, "Cornell"),
        ("Syracuse", args.syracuse_csv, "Syracuse"),
        ("Slade", args.slade_csv, "Slade"),
        ("Yale", args.yale_csv, "Yale"),
        ("CMU", args.cmu_csv, "CMU"),
        ("GSA", args.gsa_csv, "GSA"),
        ("Kunstakademie Dusseldorf", args.duesseldorf_csv, "Kunstakademie Dusseldorf"),
    ]

    for sheet_title, csv_path, school_label in sources:
        if not os.path.exists(csv_path):
            print(f"Skipping {sheet_title}: {csv_path} not found")
            continue
        rows = read_csv_rows(csv_path)
        write_sheet(wb, sheet_title, rows, school_label)
        print(f"{sheet_title}: {len(rows)} faculty -> sheet '{sheet_title[:31]}'")

    wb.save(args.out)
    print(f"\nSaved master faculty workbook -> {args.out}")


if __name__ == "__main__":
    main()
