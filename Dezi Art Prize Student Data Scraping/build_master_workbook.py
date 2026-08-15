import argparse
import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

UNIFIED_COLUMNS = [
    "name",
    "email",
    "major",
    "graduation_year",
    "portfolio_url",
    "college",
    "contacted",
    "response",
    "notes",
    "sent_to_nitya",
]

# (sheet title, csv filename, college label). Single shared source list -
# both split_master_workbook.py and build_nitya_batch.py import this rather
# than keeping their own copy, since a second hardcoded copy previously
# drifted out of sync and silently dropped Cranbrook from one output.
SOURCES = [
    ("RIT", "rit_students.csv", "RIT"),
    ("MCAD", "mcad_students.csv", "MCAD"),
    ("Cranbrook", "cranbrook_students.csv", "Cranbrook"),
    ("RISD", "risd_students.csv", "RISD"),
    ("Otis", "otis_students.csv", "Otis"),
    ("Parsons", "parsons_students.csv", "Parsons"),
    ("Temple-Tyler", "temple_students.csv", "Temple/Tyler"),
    ("VCU", "vcu_students.csv", "VCU"),
    ("Yale", "yale_students.csv", "Yale"),
    ("CMU", "cmu_students.csv", "CMU"),
    ("UW-Madison", "uw_madison_students.csv", "UW-Madison"),
    ("Ohio State", "ohio_state_students.csv", "Ohio State"),
    ("Pratt", "pratt_students.csv", "Pratt"),
    ("BU", "bu_students.csv", "BU"),
    ("SVA", "sva_students.csv", "SVA"),
    ("MICA", "mica_students.csv", "MICA"),
    ("CCA", "cca_students.csv", "CCA"),
    ("MassArt", "massart_students.csv", "MassArt"),
    ("U Michigan Stamps", "umich_students.csv", "U Michigan Stamps"),
    ("CalArts", "calarts_students.csv", "CalArts"),
    ("UCLA", "ucla_students.csv", "UCLA"),
    ("Columbia College Chicago", "columbia_students.csv", "Columbia College Chicago"),
    ("Iowa", "iowa_students.csv", "University of Iowa"),
    ("UW Art", "uw_art_students.csv", "University of Washington"),
    ("BGSU", "bgsu_students.csv", "BGSU"),
    ("CalArts Animation", "calarts_animation_students.csv", "CalArts (Animation)"),
    ("UAL Central Saint Martins", "ual_csm_students.csv", "UAL Central Saint Martins"),
    ("Alfred University", "alfred_students.csv", "Alfred University"),
    ("Cornell", "cornell_students.csv", "Cornell"),
    ("UT Austin", "ut_austin_students.csv", "UT Austin"),
    ("Cooper Union", "cooper_union_students.csv", "Cooper Union"),
    ("SCAD", "scad_students.csv", "SCAD"),
]

RESPONSE_OPTIONS = "No response,Interested,Not interested,Bounced,Follow-up needed"

# One citation per school: where its data was publicly listed, so outreach can
# credibly point to the exact source. Update this whenever a school's scraper
# is added or changes source pages. List multiple sources semicolon-separated
# if a school's data was aggregated from more than one page (e.g. MCAD).
SOURCE_CITATIONS = {
    "RIT": {
        "name": "RIT Creativity Portal",
        "url": "https://creativity.cad.rit.edu/",
        "accessed": "2026-07-31",
    },
    "MCAD": {
        "name": "MCAD Spring 2022 Commencement Page; MCAD Spring 2023 Commencement Ceremony Page",
        "url": "https://www.mcad.edu/events/spring-2022-commencement; "
               "https://www.mcad.edu/events/spring-2023-commencement-ceremony",
        "accessed": "2026-07-31",
    },
    "Cranbrook": {
        "name": "Cranbrook Academy of Art Alumni Directory — 2026 only, filtered to "
                "Painting/Photography/Fiber/Sculpture/Ceramics/Graphic Design/2D/3D/4D "
                "Design (Print Media, Metalsmithing, Architecture, Industrial Design "
                "excluded per user request)",
        "url": "https://cranbrookart.edu/alumni/directory/",
        "accessed": "2026-07-31",
    },
    "RISD": {
        "name": "RISD Grad Show 2026 Student Index",
        "url": "https://publications.risd.edu/grad-show-2026/student-index",
        "accessed": "2026-07-31",
    },
    "Otis": {
        "name": "Otis College 2026 Annual Exhibition — All Students",
        "url": "https://www.otis.edu/about/our-work/annual-exhibition/2026/all-students.html",
        "accessed": "2026-07-31",
    },
    "Parsons": {
        "name": "Parsons Fine Arts Thesis pages — 2025 BFA Thesis, 2025 MFA Thesis, "
                "2026 MFA Thesis",
        "url": "https://amt.parsons.edu/finearts/2025-bfa-thesis/; "
               "https://amt.parsons.edu/finearts/2025-mfa-thesis/; "
               "https://amt.parsons.edu/finearts/2026-mfa-thesis/",
        "accessed": "2026-08-01",
    },
    "Temple/Tyler": {
        "name": "Temple University Tyler School of Art — 2025 MFA Thesis Exhibitions page "
                "(names only, from exhibition schedule text — no email/major/portfolio "
                "published on source page)",
        "url": "https://tyler.temple.edu/2025-mfa-thesis-exhibitions-rewoven-collective-stories",
        "accessed": "2026-08-01",
    },
    "VCU": {
        "name": "VCU ICA (Institute for Contemporary Art) 2025 MFA Thesis Exhibition page "
                "(names only, from artwork photo caption credits — likely incomplete vs. "
                "full graduating cohort)",
        "url": "https://icavcu.org/exhibitions/2025-mfa-thesis/",
        "accessed": "2026-08-01",
    },
    "Yale": {
        "name": "Yale School of Art exhibition pages — Painting/Printmaking, Sculpture, "
                "Photography, and Graphic Design MFA Thesis, 2025 and 2026 (names only, "
                "from exhibition 'Featuring' text — no email/portfolio published)",
        "url": "https://art.yale.edu/exhibitions/spring-2025-painting-thesis (and 7 "
               "corresponding spring-{2025,2026}-{painting,sculpture,photography,"
               "graphic-design}-thesis pages)",
        "accessed": "2026-08-01",
    },
    "CMU": {
        "name": "Carnegie Mellon School of Art — MFA Students directory + per-student "
                "profile pages",
        "url": "https://art.cmu.edu/mfa/students/",
        "accessed": "2026-08-01",
    },
    "UW-Madison": {
        "name": "UW-Madison Department of Art — Graduate Students directory",
        "url": "https://art.wisc.edu/people/graduate-students/",
        "accessed": "2026-08-01",
    },
    "Ohio State": {
        "name": "Ohio State Department of Art — 2025 'Desire Lines' MFA Thesis Exhibition "
                "page (clean list) and 2026 'Waiting for the Light to Change' MFA Thesis "
                "Exhibition page (names UNVERIFIED — source lists them as ambiguous "
                "run-on text, reconstructed best-effort, see notes column)",
        "url": "https://art.osu.edu/events/mfa-thesis-exhibition-desire-lines; "
               "https://uas.osu.edu/events/waiting-light-change",
        "accessed": "2026-08-01",
    },
    "Pratt": {
        "name": "Pratt Institute 2026 'Pratt Shows' pages — MFA Thesis Parts 1-2, BFA "
                "Painting/Drawing/Sculpture/Printmaking (clean 'Exhibiting Artists' lists), "
                "and BFA/MFA Photography shows (names UNVERIFIED — source lists them as "
                "ambiguous <br>-separated chunks, see notes column)",
        "url": "https://www.pratt.edu/events/pratt-shows-mfa-thesis-exhibition-part-1/ "
               "(and 23 corresponding pratt-shows-* discipline/photography pages)",
        "accessed": "2026-08-01",
    },
    "BU": {
        "name": "Boston University College of Fine Arts — 2024 MFA Thesis 'Exhibiting "
                "Students by Program' page (Painting, Sculpture, Visual Narrative, Print "
                "Media & Photography, Graphic Design)",
        "url": "https://www.bu.edu/cfa/featured-work/mfa-thesis-2024/",
        "accessed": "2026-08-01",
    },
    "SVA": {
        "name": "SVA Events & Exhibitions search (2026-dated exhibitions with an "
                "'Exhibiting artists include' / 'Artists include' description) — names "
                "only; these exhibitions may mix current students, alumni, faculty, and "
                "guest artists together, not confirmed to be current students only",
        "url": "https://sva.edu/events/search/type/Exhibition (22 paginated list pages, "
               "individual event pages fetched for all 2026-dated exhibitions)",
        "accessed": "2026-08-01",
    },
    "MICA": {
        "name": "MICA Grad Show 2026 — 'Participating students' per-program pages "
                "(14 MA/MFA programs), each with name and website/Instagram link",
        "url": "https://www.mica.edu/events-exhibitions/annual-events-series/commencement/"
               "grad-show-2026/ (14 program subpages)",
        "accessed": "2026-08-02",
    },
    "CCA": {
        "name": "California College of the Arts newsroom article announcing the 2026 "
                "MFA Fine Arts Thesis Exhibition (names only; the actual event pages on "
                "portal.cca.edu are behind a student/staff login wall and were not used)",
        "url": "https://www.cca.edu/newsroom/cca-presents-the-2026-mfa-fine-arts-graduate-exhibitions/",
        "accessed": "2026-08-02",
    },
    "MassArt": {
        "name": "MassArt 2026 MFA Thesis Exhibition, Parts I & II — 'FEATURED ARTISTS' "
                "lists with name + program (names only, no email/portfolio)",
        "url": "https://calendar.massart.edu/event/2026-mfa-thesis-exhibition-PARTI; "
               "https://calendar.massart.edu/event/2026-spring-mfa-thesis-exhibition-part-ii",
        "accessed": "2026-08-02",
    },
    "U Michigan Stamps": {
        "name": "U Michigan Stamps 2025 MFA Thesis Exhibition page (clean, dated) plus "
                "the Graduate and Undergraduate Research & Creative Work gallery pages "
                "(mixed: some entries have a reliable name+year, many are UNVERIFIED "
                "'Name: Artwork Title' entries with no year and unconfirmed name/title "
                "order — see notes column per row). Site is behind a Cloudflare JS "
                "challenge; fetched via crawl4ai headless browser, not a plain HTTP fetch.",
        "url": "https://stamps.umich.edu/events/2025-mfa-thesis-exhibition; "
               "https://stamps.umich.edu/research-creative-work/graduate-work-mfa; "
               "https://stamps.umich.edu/research-creative-work/undergraduate-work",
        "accessed": "2026-08-02",
    },
    "CalArts": {
        "name": "CalArts 'High Pass' BFA Class of 2026 group exhibition page (names "
                "only; site requires crawl4ai due to cookie-consent/Cloudflare Turnstile "
                "blocking a plain fetch)",
        "url": "https://calarts.edu/high-pass",
        "accessed": "2026-08-02",
    },
    "UCLA": {
        "name": "UCLA Department of Art — current Graduate Students directory by area "
                "of study (Ceramics, Interdisciplinary Studio, New Genres, Painting and "
                "Drawing, Photography, Sculpture), each linking to a personal "
                "portfolio/Instagram where available",
        "url": "https://www.art.ucla.edu/graduate-students",
        "accessed": "2026-08-02",
    },
    "Columbia College Chicago": {
        "name": "Columbia College Chicago — 'Human Condition' 2026 BA/BFA Fine Art "
                "Exhibition page, Hokin Gallery and C33 Gallery exhibitor lists (names "
                "only). Note: 2 of 24 names have unusual letter-spacing in the source "
                "HTML itself (e.g. 'Faith H o g a n') — kept verbatim, flagged in notes.",
        "url": "https://students.colum.edu/ssac/exhibition-archives/Manifest-Exhibitions/"
               "2026/human-condition-2026-babfa-in-fine-art-exhibition",
        "accessed": "2026-08-02",
    },
    "University of Iowa": {
        "name": "University of Iowa School of Art, Art History, and Design — MFA "
                "Virtual Exhibitions page (2024-2026 cohorts), each linking to a "
                "Matterport 3D exhibition tour rather than a personal portfolio site",
        "url": "https://art.uiowa.edu/events/mfa-virtual-exhibitions",
        "accessed": "2026-08-04",
    },
    "University of Washington": {
        "name": "University of Washington 2026 MFA + MDes Thesis Exhibition page "
                "(Henry Art Gallery) — names only; source does not map a specific "
                "discipline to each individual student",
        "url": "https://henryart.org/exhibitions/2026-university-of-washington-mfa-mdes-thesis-exhibition",
        "accessed": "2026-08-04",
    },
    "BGSU": {
        "name": "BGSU ScholarWorks — School of Art MA and MFA Graduate Exhibitions, "
                "Portfolios, and Theses repository, 2025 cohort (thesis title + "
                "author name; portfolio_url is the repository record, not a personal "
                "site)",
        "url": "https://scholarworks.bgsu.edu/ms_art/",
        "accessed": "2026-08-04",
    },
    "CalArts (Animation)": {
        "name": "CalArts Film/Video Animation Student Portfolios, 2026 — Character "
                "Animation (BFA 1-4, Affiliates, Recent Alumni) and Experimental "
                "Animation (BFA 1-4, MFA 1-3, Recent Alumni), 14 pages total. Each "
                "student entry includes name, class year, specialization, and "
                "Resume/Email/Portfolio/social links in plain static HTML (no JS "
                "rendering needed). Distinct from the separate 'CalArts' source "
                "(High Pass BFA exhibition, names only, no email).",
        "url": "https://calarts.edu/filmvideo/animation-student-portfolios/2026/"
               "character-animation and .../experimental-animation (14 sub-pages: "
               "bfa-1 through bfa-4, affiliates, recent-alumni, mfa-1 through mfa-3)",
        "accessed": "2026-08-07",
    },
    "UAL Central Saint Martins": {
        "name": "UAL Showcase — Central Saint Martins graduate showcase listing "
                "(JS-rendered, scroll-triggered grid; pages 17 through end, "
                "start_rank 193-790), filtered to students whose course matched "
                "an in-scope medium. Each qualifying student's individual UAL "
                "Showcase profile page was fetched for a direct mailto: email "
                "link. Industrial Design, Curation/Culture/Criticism, and "
                "Architecture courses excluded per user decision 2026-08-15.",
        "url": "https://ualshowcase.arts.ac.uk/c/central-saint-martins"
               "?collection=ual~sp-showcase&meta_college_sand=Central+Saint+Martins "
               "(paginated, num_ranks=12, start_rank increments of 12)",
        "accessed": "2026-08-15",
    },
    "Alfred University": {
        "name": "Alfred University press releases and event pages naming 2026 and 2025 "
                "MFA Thesis Shows (Ceramic Art, Painting, Sculpture-Dimensional Studies, "
                "Electronic Integrated Arts). The original per-student thesis subpage "
                "system this project's source research relied on "
                "(thesis-exhibits/{year}/mfa/{lastname}/) has since been retired — every "
                "such URL now redirects to a generic photo-gallery hub with no "
                "consolidated roster, so this is a known-partial cohort pieced together "
                "from press releases and search-indexed event pages, not a single clean "
                "source.",
        "url": "https://www.alfred.edu/about/news/pressreleases/general/2026/"
               "nyscc-at-alfred-university-set-to-host-2026-mfa-thesis-shows.cfm "
               "(and 1 more 2026 press release, plus the 2025 MFA Thesis Exhibitions "
               "event calendar page)",
        "accessed": "2026-08-15",
    },
    "Cornell": {
        "name": "Cornell AAP (College of Architecture, Art, and Planning) event pages for "
                "several small named MFA/BFA thesis group shows — Cornell's Department of "
                "Art does not publish one consolidated roster; cohorts are split across "
                "many 5-10 person exhibitions (MFA Image Text 'SELF/ASSEMBLY', MFA "
                "Creative Visual Arts, BFA 'Thesis Group Exhibition', 'Space of "
                "Becoming', 'Senior B.F.A. Thesis Exhibition', 'with a trace'). 2 more "
                "2026 BFA sub-shows (FLEXION, 'North, South, East, West') referenced in "
                "search results but their pages were never located, so those names are a "
                "known gap.",
        "url": "https://aap.cornell.edu/events/beyond-aap/m-f-a-image-text-26-thesis-show-"
               "self-assembly/ (and 5 more aap.cornell.edu event pages)",
        "accessed": "2026-08-15",
    },
    "UT Austin": {
        "name": "UT Austin Visual Arts Center (utvac.org) event pages for the Studio Art "
                "MFA Thesis Exhibitions (2025 'Acceleration Without Arrival', 2026 'Half "
                "a Second or Less') and the 2026 Senior Art Exhibition ('Proof of Life') "
                "— each a single clean, complete roster page, the highest-yield source "
                "among the 5 schools added in this pass. A separate 2026 Design MFA show "
                "('Known Otherwise', School of Design and Creative Technologies) was "
                "deliberately excluded as a different department outside this project's "
                "medium scope.",
        "url": "https://utvac.org/event/proof-life-2026-senior-art-exhibition (and 2 more "
               "utvac.org event pages)",
        "accessed": "2026-08-15",
    },
    "Cooper Union": {
        "name": "Cooper Union 'Class of 2026: In Their Own Words' feature article — the "
                "only 2 students explicitly confirmed as School of Art via an 'A'26' "
                "suffix. Cooper Union's annual End of Year Show has no roster page at "
                "all (names attributed by tradition only) and its opening-night photo "
                "gallery mixes all 3 schools (Architecture/Art/Engineering) together "
                "with no reliable way to isolate Art students without an explicit "
                "suffix — ~20 other names seen in photo captions were excluded as "
                "either confirmed non-Art or unverifiable.",
        "url": "https://cooper.edu/about/news/class-2026-their-own-words",
        "accessed": "2026-08-15",
    },
    "SCAD": {
        "name": "SCAD Thesis Digital Collection (library.scad.edu), a static/plain-HTML "
                "library catalog covering 50+ programs since Fall 2010 (MFA) / Spring "
                "2020 (undergrad) — scoped to 2025-2026 only, matching this project's "
                "convention for every other school (current/recent cohorts, not full "
                "historical archives). 11 of SCAD's 50+ programs map to this project's "
                "in-scope mediums: Painting, Illustration, Printmaking (-> Painting/"
                "Drawing), Sculpture, Fibers (-> Fiber and Material Arts), Photography, "
                "Animation (-> 2D/3D Animation), Film & Television (-> Filmmaking), "
                "Graphic Design (-> Design), User Experience (UX) Design (-> UI/UX "
                "Design), Fashion. Scraped via a purpose-built script "
                "(scad_scraper.py) since the catalog is far too large to browse "
                "manually (486 total Painting results alone, back to 2010).",
        "url": "https://library.scad.edu/screens/theses.html (11 saved-search URLs, "
               "one per in-scope program)",
        "accessed": "2026-08-15",
    },
}


def read_csv_rows(csv_path):
    with open(csv_path, "r", newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def write_sheet(wb, sheet_title, rows, college_name):
    ws = wb.create_sheet(title=sheet_title[:31])

    citation = SOURCE_CITATIONS.get(college_name)
    header_row_offset = 0
    if citation:
        ws.append([f"Source: {citation['name']}"])
        ws.append([f"URL: {citation['url']}"])
        ws.append([f"Accessed: {citation['accessed']}"])
        for r in range(1, 4):
            ws.cell(row=r, column=1).font = Font(italic=True, size=9, color="666666")
        ws.append([])
        header_row_offset = 4

    header_row_num = header_row_offset + 1
    ws.append(UNIFIED_COLUMNS)
    for cell in ws[header_row_num]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([
            row.get("name", ""),
            row.get("email", ""),
            row.get("major", ""),
            row.get("graduation_year", ""),
            row.get("portfolio_url", ""),
            row.get("college", college_name),
            row.get("contacted", "No") or "No",
            row.get("response", ""),
            row.get("notes", ""),
            row.get("sent_to_nitya", ""),
        ])

    last_row = ws.max_row
    if last_row <= header_row_num:
        last_row = header_row_num + 1
    first_data_row = header_row_num + 1

    contacted_col = UNIFIED_COLUMNS.index("contacted") + 1
    contacted_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True, showDropDown=False)
    contacted_dv.add(f"{get_column_letter(contacted_col)}{first_data_row}:{get_column_letter(contacted_col)}{last_row}")
    ws.add_data_validation(contacted_dv)

    response_col = UNIFIED_COLUMNS.index("response") + 1
    response_dv = DataValidation(type="list", formula1=f'"{RESPONSE_OPTIONS}"', allow_blank=True, showDropDown=False)
    response_dv.add(f"{get_column_letter(response_col)}{first_data_row}:{get_column_letter(response_col)}{last_row}")
    ws.add_data_validation(response_dv)

    ws.freeze_panes = f"A{first_data_row}"
    ws.auto_filter.ref = f"A{header_row_num}:{get_column_letter(len(UNIFIED_COLUMNS))}{last_row}"

    widths = {"name": 24, "email": 28, "major": 26, "graduation_year": 16,
              "portfolio_url": 36, "college": 14, "contacted": 12, "response": 18,
              "notes": 40, "sent_to_nitya": 16}
    for i, col in enumerate(UNIFIED_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 18)

    return ws


def main():
    parser = argparse.ArgumentParser(description="Build/update the master student outreach workbook")
    parser.add_argument("--out", default="master_students.xlsx")
    args = parser.parse_args()

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_title, csv_path, college_name in SOURCES:
        if not os.path.exists(csv_path):
            print(f"Skipping {sheet_title}: {csv_path} not found")
            continue
        rows = read_csv_rows(csv_path)
        write_sheet(wb, sheet_title, rows, college_name)
        print(f"{sheet_title}: {len(rows)} students -> sheet '{sheet_title[:31]}'")

    wb.save(args.out)
    print(f"\nSaved master workbook -> {args.out}")


if __name__ == "__main__":
    main()
